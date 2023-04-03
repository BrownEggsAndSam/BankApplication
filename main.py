import openpyxl
from openpyxl import Workbook

def create_excel_report(general_report, tables_columns):
    wb = Workbook()

    # Create a general report sheet
    general_sheet = wb.active
    general_sheet.title = "General Report"

    for idx, line in enumerate(general_report.split('\n'), start=1):
        general_sheet.cell(row=idx, column=1, value=line)

    # Create a sheet for each table
    for table, columns in tables_columns.items():
        table_sheet = wb.create_sheet(title=table)
        table_sheet.cell(row=1, column=1, value=f"Table: {table}")
        table_sheet.cell(row=2, column=1, value="Columns used:")

        for idx, column in enumerate(columns, start=1):
            table_sheet.cell(row=idx+2, column=2, value=column)

    # Save the workbook
    wb.save("report.xlsx")

# Generate a general report
general_report = ""
general_report += f"Tables used: {', '.join(tables_columns.keys())}\n"

if join_type and join_condition:
    general_report += "Join operation:\n"
    general_report += f"  - Type: {join_type}\n"
    general_report += f"  - Join condition: {join_condition}\n"

if parsed.token_first(skip_ws=True, skip_cm=True).value.upper() == "SELECT":
    orderby_found = False
    for token in parsed.tokens:
        if token.ttype == Keyword and token.value.upper() == "ORDER":
            orderby_found = True
        elif orderby_found and token.ttype == Keyword and token.value.upper() == "BY":
            order_columns = [str(t) for t in token.parent.tokens[token.parent.token_index(token) + 1:][0].get_identifiers()]
            general_report += "Ordering:\n"
            general_report += f"  - Columns: {', '.join(order_columns)}\n"
            break

# Create and save the Excel report
create_excel_report(general_report, tables_columns)
